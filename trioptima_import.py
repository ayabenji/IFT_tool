from __future__ import annotations

import importlib
from datetime import date
from pathlib import Path
from typing import Dict, List, Tuple

try:  # Pandas peut être absent au démarrage (chargement paresseux)
    import pandas as pd
except ModuleNotFoundError:  # pragma: no cover - fallback runtime
    pd = None  # type: ignore[assignment]

from openpyxl import load_workbook

from excel_read import _norm
from template_write import build_targets_index, letter_to_index

__all__ = [
    "expected_trioptima_prefix",
    "locate_trioptima_file",
    "load_trioptima_table",
    "aggregate_trioptima",
    "build_trioptima_mapping",
    "apply_trioptima_to_workbook",
    "filter_bndfwd_rows",
    "apply_bndfwd_to_workbook",

]


def _ensure_pandas() -> "pd":
    global pd
    if pd is None:
        pd = importlib.import_module("pandas")  # type: ignore[assignment]
    return pd  # type: ignore[return-value]


def expected_trioptima_prefix(ifts_date: date) -> str:
    """Préfixe attendu pour le fichier TriOptima à partir de la date de prod."""

    return f"search_groupama-am_{ifts_date:%Y-%m-%d}"


def locate_trioptima_file(prod_dir: Path, ifts_date: date) -> Path:
    """Localise le fichier TriOptima à partir du préfixe attendu."""

    prefix = expected_trioptima_prefix(ifts_date)
    direct_pattern = f"{prefix}*.csv"
    matches = sorted(prod_dir.glob(direct_pattern))
    if matches:
        return matches[0]

    prefix_lower = prefix.lower()
    fallback = [
        path
        for path in sorted(prod_dir.glob("*.csv"))
        if path.name.lower().startswith(prefix_lower)
    ]
    if fallback:
        return fallback[0]

    raise FileNotFoundError(
        f"Fichier TriOptima introuvable dans {prod_dir}: {direct_pattern}"
    )


def _to_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\u00a0", " ")
    if not text:
        return None
    text = text.replace(" ", "").replace("'", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _sub_optional(left: object, right: object) -> float | None:
    lf = _to_float(left)
    rf = _to_float(right)
    if lf is None and rf is None:
        return None
    return (lf or 0.0) - (rf or 0.0)


def _div_optional(numerator: object, denominator: object) -> float | None:
    num = _to_float(numerator)
    den = _to_float(denominator)
    if num is None or den is None or den == 0:
        return None
    return num / den


def load_trioptima_table(path: Path) -> "pd.DataFrame":
    """Charge le CSV TriOptima et prépare les colonnes nécessaires."""

    _ensure_pandas()
    df = pd.read_csv(path,decimal='.',sep=';')  # type: ignore[attr-defined]

    required = {"FREE_TEXT_2", "MTM_VALUE", "MTM_DIFF"}
    missing = required.difference(df.columns)
    if missing:
        raise KeyError(
            f"Colonnes manquantes dans {path.name}: {', '.join(sorted(missing))}"
        )

    df = df.copy()

    def _clean_free_text(value: object) -> str:
        if pd.isna(value):  # type: ignore[arg-type]
            return ""
        return str(value).strip()

    df["FREE_TEXT_2"] = df["FREE_TEXT_2"].apply(_clean_free_text)

    codes = df["FREE_TEXT_2"].str.split("/", n=1).str[0].str.strip()
    df["Code DI"] = codes.fillna("")

    extras = {"FREE_TEXT_1", "BOOK", "CP", "NOTIONAL"}
    present_extras = extras.intersection(df.columns)
    missing_extras = extras.difference(df.columns)
    for column in sorted(missing_extras):
        df[column] = None
    df.attrs["missing_bndfwd_columns"] = sorted(missing_extras)

    def _clean_str(value: object) -> str:
        if pd.isna(value):  # type: ignore[arg-type]
            return ""
        return str(value).strip()

    for column in ("FREE_TEXT_1", "CP"):
        if column in present_extras:
            df[column] = df[column].apply(_clean_str)

    if "BOOK" in present_extras:
        df["BOOK"] = df["BOOK"].apply(_clean_str)

    for column in ("MTM_VALUE", "MTM_DIFF", "NOTIONAL"):
        if column in df.columns:
            df[column] = df[column].apply(_to_float)
            df[column] = pd.to_numeric(df[column], errors="coerce")  # type: ignore[attr-defined]

    if {"MTM_VALUE", "MTM_DIFF"}.issubset(df.columns):
        df["MTM_CONTREPARTIE"] = df["MTM_VALUE"] - df["MTM_DIFF"]

    keep = [
        "Code DI",
        "FREE_TEXT_1",
        "BOOK",
        "CP",
        "NOTIONAL",
        "MTM_VALUE",
        "MTM_DIFF",
        "MTM_CONTREPARTIE",
    ]
    existing = [col for col in keep if col in df.columns]
    return df[existing]


def _normalize_book(value: object) -> str | None:
    num = _to_float(value)
    if num is not None:
        if float(num).is_integer():
            return str(int(num))
        return str(num)
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.lower() == "nan":
        return None
    return text


def filter_bndfwd_rows(df: "pd.DataFrame") -> "pd.DataFrame":
    """Filtre les lignes TriOptima pertinentes pour la feuille BND FWD."""

    _ensure_pandas()
    if df.empty:
        return df.copy()
    
    required = {"FREE_TEXT_1", "BOOK", "CP", "NOTIONAL", "MTM_VALUE", "MTM_DIFF"}
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise KeyError(
            "Colonnes manquantes pour le traitement BND FWD: "
            + ", ".join(sorted(missing))
        )

    filtered = df.copy()
    filtered["FREE_TEXT_1"] = (
        filtered["FREE_TEXT_1"].astype(str).str.strip().str.replace("\u00a0", " ")
    )
    mask_prefix = filtered["FREE_TEXT_1"].str.upper().str.startswith("BDFWD")
    
    filtered["BOOK"] = filtered["BOOK"].apply(_normalize_book)
    allowed_books = {"601", "602", "603"}
    mask_book = filtered["BOOK"].isin(allowed_books)

    result = filtered[mask_prefix & mask_book].copy()
    result["CP"] = result["CP"].apply(lambda v: str(v).strip() if not pd.isna(v) else "")  # type: ignore[arg-type]

    numeric_cols = ["NOTIONAL", "MTM_VALUE", "MTM_DIFF", "MTM_CONTREPARTIE"]
    for column in numeric_cols:
        if column in result.columns:
            result[column] = pd.to_numeric(result[column], errors="coerce")  # type: ignore[attr-defined]

    return result.reset_index(drop=True)


   

def aggregate_trioptima(df: "pd.DataFrame") -> "pd.DataFrame":
    """Agrège les montants TriOptima par Code DI."""

    _ensure_pandas()
    if df.empty:
        return df.copy()
    
    if "Code DI" not in df.columns:
        raise KeyError("Colonne 'Code DI' manquante pour l'agrégation TriOptima")

    df = df.copy()

    df["Code DI"] = df["Code DI"].apply(
        lambda v: str(v).strip() if not pd.isna(v) else ""
    )  # type: ignore[arg-type]
    df = df[df["Code DI"] != ""]

    numeric_cols = ["MTM_VALUE", "MTM_DIFF", "MTM_CONTREPARTIE"]
    
    existing_numeric = [col for col in numeric_cols if col in df.columns]

    if df.empty:
        return pd.DataFrame(columns=["Code DI", *existing_numeric])

    for column in existing_numeric:
        df[column] = df[column].fillna(0.0)
    grouped = (
        df.groupby("Code DI", as_index=False)[existing_numeric]
        .sum()
        .sort_values("Code DI")
        .reset_index(drop=True)
    )
    return grouped


def build_trioptima_mapping(df: "pd.DataFrame") -> Dict[str, float]:
    """Construit un dictionnaire Code DI → MTM contrepartie agrégé."""

    mapping: Dict[str, float] = {}
    _ensure_pandas()
    if df.empty:
        return mapping
    for _, row in df.iterrows():
        code = str(row["Code DI"]).strip()
        if not code:
            continue
        value = row.get("MTM_CONTREPARTIE")
        if pd.isna(value):  # type: ignore[arg-type]
            continue
        value_float = _to_float(value)
        if value_float is None:
            continue
        mapping[code] = value_float
    return mapping


def apply_trioptima_to_workbook(
    workbook_path: Path,
    mtm_mapping: Dict[str, float],
    *,
    header_row: int = 6,
    sheet: str = "IRS - INF – XCCY",
) -> Tuple[int, List[str], List[dict[str, object]], List[str]]:
    """Injecte les MTM contrepartie dans le template et recalcule les écarts."""

    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    if not mtm_mapping:
        return 0, [], [], []

    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    wb = load_workbook(workbook_path, keep_vba=keep_vba, data_only=False)
    if sheet not in wb.sheetnames:
        raise KeyError(f"Feuille '{sheet}' absente de {workbook_path.name}")
    ws = wb[sheet]

    targets = build_targets_index(ws, header_row)
    code_cols = targets.get(_norm("Code DI"), [])
    if not code_cols:
        raise KeyError("Colonne 'Code DI' introuvable dans le template")
    code_col = code_cols[0]

    start_row = header_row + 1
    updated = 0
    missing_codes: List[str] = []
    used_codes: set[str] = set()
    preview: List[dict[str, object]] = []

    col_notional = letter_to_index("M") + 1

    def _cell_value(letter: str, row_idx: int) -> object:
        idx = letter_to_index(letter) + 1
        return ws.cell(row=row_idx, column=idx).value

    def _set_value(letter: str, row_idx: int, value: object) -> None:
        idx = letter_to_index(letter) + 1
        ws.cell(row=row_idx, column=idx).value = value

    for row in range(start_row, ws.max_row + 1):
        code_cell = ws.cell(row=row, column=code_col).value
        if code_cell is None or (isinstance(code_cell, str) and not code_cell.strip()):
            break
        code = str(code_cell).strip()
        mtm_value = mtm_mapping.get(code)
        if mtm_value is None:
            missing_codes.append(code)
            continue

        used_codes.add(code)

        an_value = _cell_value("AN", row)
        notional = ws.cell(row=row, column=col_notional).value

        bd_value = _sub_optional(an_value, mtm_value)
        be_value = _div_optional(bd_value, notional)
        ax_value = _div_optional(mtm_value, notional)

        _set_value("AW", row, mtm_value)
        _set_value("AX", row, ax_value)
        _set_value("BD", row, bd_value)
        _set_value("BE", row, be_value)

        # Recalcule les colonnes de comparaison dépendantes
        _set_value("BK", row, _sub_optional(an_value, mtm_value))
        _set_value("BS", row, _sub_optional(an_value, _cell_value("BD", row)))
        _set_value("BT", row, _sub_optional(_cell_value("AO", row), _cell_value("BE", row)))
        _set_value(
            "CA",
            row,
            _sub_optional(_cell_value("AW", row), _cell_value("BD", row)),
        )
        _set_value(
            "CB",
            row,
            _sub_optional(_cell_value("AX", row), _cell_value("BE", row)),
        )

        updated += 1
        preview.append(
            {
                "Code DI": code,
                "MTM Contrepartie": mtm_value,
                "Diff (AN-AW)": bd_value,
                "Diff / Notional": be_value,
            }
        )

    wb.save(workbook_path)

    unused_codes = sorted(set(mtm_mapping).difference(used_codes))
    return updated, missing_codes, preview, unused_codes

def apply_bndfwd_to_workbook(
    workbook_path: Path,
    rows: "pd.DataFrame",
    *,
    sheet: str = "BND FWD",
    start_row: int = 2,
    threshold: float = 0.005,
) -> Tuple[int, List[str], List[dict[str, object]], List[str]]:
    """Injecte les lignes BND FWD filtrées dans le template."""

    _ensure_pandas()
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)

    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    wb = load_workbook(workbook_path, keep_vba=keep_vba, data_only=False)
    if sheet not in wb.sheetnames:
        raise KeyError(f"Feuille '{sheet}' absente de {workbook_path.name}")

    ws = wb[sheet]

    records = rows.to_dict(orient="records") if not rows.empty else []

    max_clear_row = max(ws.max_row, start_row + max(len(records), 1))
    last_col_idx = letter_to_index("L") + 1
    for row_idx in range(start_row, max_clear_row + 1):
        for col_idx in range(1, last_col_idx + 1):
            ws.cell(row=row_idx, column=col_idx).value = None

    updated = 0
    missing_data: List[str] = []
    preview: List[dict[str, object]] = []
    alerts: List[str] = []

    alert_col = letter_to_index("K") + 1

    current_row = start_row
    for idx, record in enumerate(records):
        free_text = str(record.get("FREE_TEXT_1") or "").strip()
        cp_value = str(record.get("CP") or "").strip()
        book_value = _normalize_book(record.get("BOOK"))
        notional_value = _to_float(record.get("NOTIONAL"))
        mtm_value = _to_float(record.get("MTM_VALUE"))
        mtm_diff = _to_float(record.get("MTM_DIFF"))
        prix_ctrp = _sub_optional(mtm_value, mtm_diff)
        mtm_ctrp = prix_ctrp
        ratio_value = _div_optional(mtm_value, notional_value)
        ratio_ctrp = _div_optional(mtm_ctrp, notional_value)
        diff_ratio = (
            None
            if ratio_value is None or ratio_ctrp is None
            else ratio_value - ratio_ctrp
        )

        missing_fields: List[str] = []
        if not free_text:
            missing_fields.append("FREE_TEXT_1")
        if not cp_value:
            missing_fields.append("CP")
        if notional_value is None or notional_value == 0:
            missing_fields.append("NOTIONAL")
        if mtm_value is None:
            missing_fields.append("MTM_VALUE")
        if mtm_diff is None:
            missing_fields.append("MTM_DIFF")

        skip_row = any(field in {"NOTIONAL", "MTM_VALUE", "MTM_DIFF"} for field in missing_fields)
        label = free_text or f"Ligne {idx + 1}"
        if skip_row:
            missing_data.append(f"{label} → données manquantes: {', '.join(missing_fields)}")
            continue

        if missing_fields:
            missing_data.append(f"{label} → données manquantes: {', '.join(missing_fields)}")

        ws.cell(row=current_row, column=letter_to_index("A") + 1).value = free_text or None
        ws.cell(row=current_row, column=letter_to_index("B") + 1).value = book_value
        ws.cell(row=current_row, column=letter_to_index("C") + 1).value = cp_value or None
        ws.cell(row=current_row, column=letter_to_index("D") + 1).value = notional_value
        ws.cell(row=current_row, column=letter_to_index("E") + 1).value = mtm_value
        ws.cell(row=current_row, column=letter_to_index("F") + 1).value = prix_ctrp
        ws.cell(row=current_row, column=letter_to_index("G") + 1).value = ratio_value
        ws.cell(row=current_row, column=letter_to_index("H") + 1).value = ratio_ctrp
        ws.cell(row=current_row, column=letter_to_index("I") + 1).value = threshold
        ws.cell(row=current_row, column=letter_to_index("J") + 1).value = diff_ratio

        is_alert = diff_ratio is not None and abs(diff_ratio) > threshold
        ws.cell(row=current_row, column=alert_col).value = "alerte" if is_alert else None
        if is_alert:
            alerts.append(label)

        preview.append(
            {
                "FREE_TEXT_1": free_text,
                "BOOK": book_value,
                "CP": cp_value,
                "Notional": notional_value,
                "MTM GAM": mtm_value,
                "Prix CTP": prix_ctrp,
                "MTM %": ratio_value,
                "CTP %": ratio_ctrp,
                "Diff": diff_ratio,
                "Alerte": "oui" if is_alert else "",
            }
        )

        current_row += 1
        updated += 1

    wb.save(workbook_path)
    return updated, missing_data, preview, alerts
