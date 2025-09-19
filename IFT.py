import streamlit as st
from pathlib import Path
from datetime import date, timedelta
import zipfile
import shutil

# -----------------------------
# Helpers
# -----------------------------
PERIMETER_XLS_PREFIXES = ("IR_", "XCY_IR")

def next_business_day(d: date) -> date:
    """Return the next business day (Mon–Fri), ignoring public holidays."""
    nd = d + timedelta(days=1)
    while nd.weekday() >= 5:  # 5=Sat, 6=Sun
        nd += timedelta(days=1)
    return nd


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def unique_path(dest_dir: Path, filename: str) -> Path:
    """Return a non-colliding path in dest_dir for filename, adding _1, _2, ... if needed."""
    base = Path(filename).stem
    ext = Path(filename).suffix
    candidate = dest_dir / filename
    i = 1
    while candidate.exists():
        candidate = dest_dir / f"{base}_{i}{ext}"
        i += 1
    return candidate


def extract_xls_from_zip(zip_path: Path, dest_dir: Path) -> list[Path]:
    """Extract only .xls files from a ZIP to dest_dir (flattened), avoid overwriting by suffixing.
    Returns list of extracted paths.
    """
    extracted: list[Path] = []
    with zipfile.ZipFile(zip_path, "r") as z:
        for info in z.infolist():
            name = info.filename
            if name.lower().endswith(".xls") and not name.endswith("/"):
                # Flatten any subfolders inside the zip
                target = unique_path(dest_dir, Path(name).name)
                with z.open(info) as src, open(target, "wb") as out:
                    shutil.copyfileobj(src, out)
                extracted.append(target)
    return extracted

def list_perimeter_xls(dest_dir: Path) -> list[Path]:
    """Return sorted .xls files limited to the expected IR perimeter prefixes."""
    prefixes = tuple(prefix.upper() for prefix in PERIMETER_XLS_PREFIXES)
    if not dest_dir.exists():
        return []
    return sorted(
        p
        for p in dest_dir.iterdir()
        if p.is_file()
        and p.suffix.lower() == ".xls"
        and p.name.upper().startswith(prefixes)
    )

# -----------------------------
# App UI
# -----------------------------
st.set_page_config(page_title="IFT Downloader", page_icon="📦", layout="centered")
st.title("📦 IFT Downloader & Extractor")

col1, col2 = st.columns(2)
with col1:
    mode = st.radio("Mode", ["Fast", "Close"], horizontal=True, index=0)
with col2:
    ifts_date = st.date_input("Date des IFTS", value=date.today())

# Derived values
file_day = next_business_day(ifts_date)
file_tag = file_day.strftime("%m%d%Y")  # e.g. 03202025

st.info(
    f"Jour ouvré utilisé pour les fichiers : **{file_day.strftime('%A %d %B %Y')}** → tag **{file_tag}**"
)

# Paths (defaults from spec)
src_dir = Path(r"S:\PRD\SuperDerivatives\In\Archives")
base_out = Path(r"C:\Users\abenjelloun\OneDrive - Cooperactions\GAM-E-Risk Perf - RMP\1.PROD\4.REPORTINGS SPEC CLIENTS\1.Groupe - IFT (CB-JB)")

year_folder = ifts_date.strftime("%Y")
month_folder = ifts_date.strftime("%m-%Y")
dest_dir = base_out / year_folder / month_folder / "prod" / mode.lower()# Search patterns
patterns = [
    f"XCY_IR_{file_tag}*.zip",
    f"IR_{file_tag}*.zip",
]

if st.button("🔎 Rechercher & extraire"):
    try:
        if not src_dir.exists():
            st.error(f"Dossier source introuvable: {src_dir}")
            st.stop()

        ensure_dir(dest_dir)

        # Find matching zip files
        found = []
        for pat in patterns:
            matches = list(src_dir.glob(pat))
            # On Windows, glob is case-insensitive; this should catch .ZIP as well
            if matches:
                st.write(f"Pattern **{pat}** → {len(matches)} fichier(s) trouvé(s)")
            else:
                st.warning(f"Pattern **{pat}** → aucun fichier trouvé")
            found.extend(matches)

        if not found:
            st.stop()

        total_extracted = []
        for zp in sorted(found):
            st.write(f"➡️ Traitement du zip : {zp.name}")
            extracted = extract_xls_from_zip(zp, dest_dir)
            if extracted:
                st.success(f"{len(extracted)} fichier(s) .xls extrait(s)")
                for p in extracted:
                    st.write(f"• {p.name}")
                total_extracted.extend(extracted)
            else:
                st.warning("Aucun .xls trouvé dans cette archive")

        if total_extracted:
            st.success(f"✅ Terminé. Fichiers extraits dans : {dest_dir}")
        else:
            st.warning("Aucun fichier .xls extrait. Vérifier les archives et le tag de date.")

    except Exception as e:
        st.exception(e)


# =============================
# Étape 2 — Périmètre & mapping de colonnes
# =============================
import pandas as pd
import numpy as np
from io import BytesIO

# --- Normalisation des noms de colonnes (robuste aux espaces/casse) ---
def _norm(s: str) -> str:
    return " ".join(str(s).strip().lower().split())

# --- Lecture XLS (smart header detection + fallback d'engine) ---
EXPECTED_HEADERS = {
    _norm(x) for x in [
        "#Ticket","Trade ID","External Id","Counterparty","Currency","Class",
        "Custom Attribute5 Value","Leg Type","Pay/Rec","Index/Fixed Rate",
        "Spread (bp)","Start Date","End Date","Notional",
        "Dirty Value","Clean Value","Accrued Interest",
    ]
}


def _score_header_row(df_nohdr: pd.DataFrame, ridx: int) -> int:
    vals = [ _norm(v) for v in df_nohdr.iloc[ridx].tolist() ]
    return sum(1 for v in vals if v in EXPECTED_HEADERS)


def _flatten_two_rows(r1: list, r2: list) -> list:
    out = []
    ln = max(len(r1), len(r2))
    for i in range(ln):
        a = str(r1[i]) if i < len(r1) and r1[i] is not None else ""
        b = str(r2[i]) if i < len(r2) and r2[i] is not None else ""
        a, b = a.strip(), b.strip()
        out.append(b if b else a)
    return out


def read_xls_smart(path: Path, search_rows: int = 12) -> pd.DataFrame:
    last_err = None
    for engine in ("calamine", "xlrd", None):
        try:
            df0 = pd.read_excel(path, sheet_name=0, header=None, engine=engine, dtype=str)
        except Exception as e:
            last_err = e
            continue
        # Chercher la meilleure ligne d'entête sur les N premières lignes
        best = (-1, -1)  # score, ridx
        for r in range(min(search_rows, len(df0))):
            s = _score_header_row(df0, r)
            if s > best[0]:
                best = (s, r)
        if best[1] == -1:
            # fallback : utiliser ligne 5 (Excel 6)
            hdr_idx = 5
            cols = df0.iloc[hdr_idx].astype(str).str.strip().tolist()
            df = df0.iloc[hdr_idx+1:].copy()
            df.columns = cols
            return df
        # Vérifier si combiner avec la ligne suivante améliore
        ridx = best[1]
        if ridx + 1 < len(df0):
            single_cols = df0.iloc[ridx].astype(str).str.strip().tolist()
            two_cols = _flatten_two_rows(df0.iloc[ridx].tolist(), df0.iloc[ridx+1].tolist())
            single_score = sum(1 for v in single_cols if _norm(v) in EXPECTED_HEADERS)
            two_score = sum(1 for v in two_cols if _norm(v) in EXPECTED_HEADERS)
            if two_score > single_score:
                cols = two_cols
                body = df0.iloc[ridx+2:].copy()
            else:
                cols = single_cols
                body = df0.iloc[ridx+1:].copy()
        else:
            cols = df0.iloc[ridx].astype(str).str.strip().tolist()
            body = df0.iloc[ridx+1:].copy()
        # Nettoyage colonnes vides nommées Unnamed... / ''
        cols = [ (c if c and not str(c).startswith("Unnamed") else "") for c in cols ]
        body.columns = cols
        # Drop colonnes complètement vides
        body = body.dropna(axis=1, how="all")
        return body
    raise RuntimeError(f"Impossible de lire {path.name}. Dernière erreur: {last_err}")

# --- Dédoublonnage/étiquetage des colonnes répétées (Leg1/Leg2/...) ---
def label_duplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    cols = list(df.columns)
    # pandas mangle: Col, Col.1, Col.2, ...
    base_counts = {}
    new_cols = []
    for c in cols:
        if isinstance(c, str) and "." in c and c.split(".")[-1].isdigit():
            base = c.rsplit(".", 1)[0]
            idx = int(c.rsplit(".", 1)[1])
            leg = idx + 1  # Col (Leg1) = base (idx missing), Col.1 (Leg2) = idx=1 → leg 2
            new_cols.append(f"{base} (Leg{leg})")
        else:
            # Si le base est déjà rencontré, on le marque (Leg1)
            base = c
            if base in base_counts:
                base_counts[base] += 1
                new_cols.append(f"{base} (Leg{base_counts[base]+1})")
            else:
                base_counts[base] = 0
                new_cols.append(base)
    df = df.copy()
    df.columns = new_cols
    return df

# --- Recherche tolérante d'une colonne par son nom logique ---
def get_col(df: pd.DataFrame, logical_name: str, required: bool = True) -> str | None:
    target = _norm(logical_name)
    mapping = {c: _norm(c) for c in df.columns}
    for c, n in mapping.items():
        if n == target:
            return c
    if required:
        raise KeyError(f"Colonne introuvable: '{logical_name}' dans {list(df.columns)[:8]}… (total {len(df.columns)})")
    return None

st.header("Étape 2 — Définir le périmètre depuis les XLS")

st.write("On charge les .xls extraits, on filtre **Custom Attribute5 Value** non vide (→ `Code DI`), et on mappe `Class` → `Classif DI` et `Class`.")

if st.button("📥 Charger & filtrer le périmètre"):
    try:
        xls_paths = list_perimeter_xls(dest_dir)
   
        if not xls_paths:
            st.warning(
                "Aucun fichier périmètre IR_*.xls ou XCY_IR_*.xls trouvé. Lance d'abord l'extraction."
            )
            st.stop()

        frames = []
        for p in xls_paths:
            st.write(f"Lecture: {p.name}")
            df = read_xls_smart(p)
            # Nettoyage basique
            df = df.dropna(how="all")
            df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
            df = label_duplicate_columns(df)
            df["__source_file__"] = p.name
            frames.append(df)

        raw = pd.concat(frames, ignore_index=True, sort=False)

        # Colonnes utiles
        col_code_di = get_col(raw, "Custom Attribute5 Value")
        col_class = get_col(raw, "Class")

        # Identifiants de secours (si présents)
        id_cols = []
        for k in ["#Ticket", "Trade ID", "External Id"]:
            try:
                id_cols.append(get_col(raw, k, required=False))
            except Exception:
                pass
        id_cols = [c for c in id_cols if c]

        # Filtre Code DI non vide
        raw[col_code_di] = raw[col_code_di].astype(str).str.strip()
        peri = raw[raw[col_code_di].notna() & (raw[col_code_di] != "")].copy()

        # Mapping de sortie
        out_cols = {
            "Code DI": col_code_di,
            "Classif DI": col_class,
            "Class": col_class,
        }
        # Garder aussi quelques méta colonnes utiles pour la suite
        keep_meta = [c for c in id_cols if c] + [get_col(raw, "Counterparty", required=False), get_col(raw, "Currency", required=False)]
        keep_meta = [c for c in keep_meta if c]

        result = pd.DataFrame({dst: peri[src] for dst, src in out_cols.items()})
        for c in keep_meta:
            result[c] = peri[c]

        # Déduplication douce par identifiants si possible
        dedup_keys = [k for k in id_cols if k]
        if dedup_keys:
            before = len(result)
            result = result.drop_duplicates(subset=dedup_keys)
            after = len(result)
            st.info(f"Déduplication: {before} → {after} lignes (clés: {dedup_keys})")

        st.success(f"✅ Périmètre construit: {len(result)} lignes")
        st.dataframe(result.head(50))

        # Sauvegarde disque + téléchargement direct
        out_name = f"perimetre_IFTS_{file_tag}.xlsx"
        out_path = dest_dir / out_name
        with pd.ExcelWriter(out_path, engine="xlsxwriter") as wr:
            result.to_excel(wr, sheet_name="Perimetre", index=False)

        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="xlsxwriter") as wr:
            result.to_excel(wr, sheet_name="Perimetre", index=False)
        st.download_button(
            label="⬇️ Télécharger le périmètre (Excel)",
            data=bio.getvalue(),
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.exception(e)

# =============================
# Helpers legs (conservés pour le support `source_leg` côté YAML)
# =============================

def _leg_token(name: str) -> str:
    n = _norm(name)
    if any(t in n for t in ["1+2", "total", "both legs", "both", "sum"]):
        return "total"
    if any(t in n for t in ["leg 1", "leg1", "(1)", " l1 "]):
        return "1"
    if any(t in n for t in ["leg 2", "leg2", "(2)", " l2 "]):
        return "2"
    return ""


def _resolve_leg_columns(df: pd.DataFrame, base: str) -> dict:
    variants = {"1": None, "2": None, "total": None}
    candidates = []
    target = _norm(base)
    for c in df.columns:
        n = _norm(str(c))
        if target == n or n.startswith(target):
            candidates.append(str(c))
    for c in candidates:
        tok = _leg_token(c)
        if tok in ("1", "2", "total") and variants[tok] is None:
            variants[tok] = c
    if variants["total"] is None:
        unsuffixed = [c for c in candidates if _leg_token(c) == ""]
        if unsuffixed:
            if (variants["1"] or variants["2"]) and len(unsuffixed) >= 1:
                variants["total"] = unsuffixed[0]
            elif len(unsuffixed) == 1 and not (variants["1"] or variants["2"]):
                variants["1"] = unsuffixed[0]
    return variants


def col_for_leg(df: pd.DataFrame, base: str, leg: int) -> str | None:
    var = _resolve_leg_columns(df, base)
    key = "1" if leg == 1 else "2"
    return var.get(key)


def col_for_total(df: pd.DataFrame, base: str) -> str | None:
    var = _resolve_leg_columns(df, base)
    return var.get("total")


# =============================
# Helpers template pour l'étape YAML
# =============================
import openpyxl
from openpyxl import load_workbook

TEMPLATE_DEFAULT = r"C:/Users/abenjelloun/OneDrive - Cooperactions/GAM-E-Risk Perf - RMP/1.PROD/4.REPORTINGS SPEC CLIENTS/1.Groupe - IFT (CB-JB)/IFT -template.xlsm"

def copy_template_to_dest(template_path: Path, dest_dir: Path, file_tag: str, mode: str) -> Path:
    ensure_dir(dest_dir)
    out_path = dest_dir / f"IFT_{file_tag}_{mode.lower()}.xlsm"
    shutil.copy2(template_path, out_path)
    return out_path

# =============================
# Étape 2.7 — Mapping YAML (source_letter + target_occurrence)
# =============================
# Permet de coller un YAML de mapping (comme celui que tu as donné) pour piloter
# l'écriture dans le template :
#  - côté source: supporte source_letter, source, source_leg
#  - côté template: target par header + target_occurrence (ou target_letter)
#  - variables: n'importe quels noms, utilisés ensuite dans computed.expr

try:
    import yaml  # type: ignore
except Exception:
    yaml = None

from openpyxl.utils.cell import column_index_from_string


def _letter_to_index(letter: str) -> int:
    """Excel letter -> 0-based index."""
    return column_index_from_string(letter) - 1


def _load_raw_with_colorders() -> tuple[pd.DataFrame, dict[str, list[str]]]:
    """Relit les .xls pour récupérer un DataFrame consolidé ET l'ordre exact des colonnes par fichier.
    Retourne (raw_df, colorders_by_file) où colorders_by_file[file_name] = [colA, colB, ...] (ordre natif).
    """
    xls_paths = list_perimeter_xls(dest_dir)
    if not xls_paths:
        raise FileNotFoundError(
            "Aucun fichier périmètre IR_*.xls ou XCY_IR_*.xls trouvé. Lance d'abord l'extraction."
        )
    frames = []
    orders: dict[str, list[str]] = {}
    for p in xls_paths:
        df = read_xls_smart(p)
        df = df.dropna(how="all")
        df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
        df = label_duplicate_columns(df)
        orders[p.name] = list(df.columns)
        df["__source_file__"] = p.name
        frames.append(df)
    raw = pd.concat(frames, ignore_index=True, sort=False)
    return raw, orders


def _value_from_source_spec(row: pd.Series, orders_by_file: dict[str, list[str]], spec: dict) -> any:
    """Lit une valeur à partir d'un spec YAML: supporte source_letter, source (nom), source_leg.
    - source_letter: utilise l'ordre de colonnes du FICHIER d'origine pour récupérer le nom de colonne et lit row[colname]
    - source: lit par nom de colonne (exact)
    - source_leg: base+leg via col_for_leg (sur le DF global)
    """
    if spec is None:
        return None
    if "source_letter" in spec and spec["source_letter"]:
        file = row.get("__source_file__")
        cols = orders_by_file.get(file, [])
        idx0 = _letter_to_index(str(spec["source_letter"]))
        if 0 <= idx0 < len(cols):
            colname = cols[idx0]
            return row.get(colname)
        return None
    if "source" in spec and spec["source"]:
        return row.get(spec["source"]) 
    if "source_leg" in spec:
        base = spec["source_leg"].get("base")
        leg  = int(spec["source_leg"].get("leg", 1))
        c = col_for_leg(pd.DataFrame([row]), base, leg)  # petite astuce locale
        # col_for_leg utilise les colonnes du DF; ici on cherche directement par nom normalisé
        # Fallback simple: essayer base (non suffixé)
        if c is None:
            c = base
        return row.get(c)
    return None


def _parse_number(x):
    try:
        if x is None:
            return None
        s = str(x).strip()
        if s == "":
            return None
        return float(s.replace(" ", "").replace(",", "."))
    except Exception:
        return None


def _build_targets_index(ws, header_row: int) -> dict[str, list[int]]:
    """Construit un index {header_norm: [col_idx1, col_idx2, ...]} pour gérer target_occurrence."""
    idx: dict[str, list[int]] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        k = _norm(str(v))
        idx.setdefault(k, []).append(c)
    return idx


def _eval_expr(expr: str, env: dict[str, float | None]) -> float | None:
    """Évalue une expression simple sur l'environnement env (variables -> nombres/None)."""
    # Remplacer None par 0 dans divisions sécurisées ? On préfère lever None si denom=0/None
    safe_env = {k: (float(v) if v is not None else None) for k, v in env.items()}
    # Helpers dans l'env
    def _div(a, b):
        if a is None or b in (None, 0):
            return None
        return a / b
    safe_env.update({"_div": _div})
    # Autoriser seulement les ops de base
    allowed = {k: v for k, v in safe_env.items()}
    allowed.update({"None": None})
    # Traduction rapide de a/b en _div(a,b) pour sécuriser les zéros (optionnel)
    # Ici on laisse l'utilisateur écrire a/b, et on gère les zéros avant.
    try:
        return eval(expr, {"__builtins__": {}}, allowed)  # nosec - environnement réduit
    except Exception:
        return None


st.header("Étape 2.7 — Coller un mapping YAML (source_letter)")

yaml_text = st.text_area("Mapping YAML", value="", height=360, help="Colle ici le YAML (source_letter/target_occurrence/variables/computed)")

colYA, colYB = st.columns(2)
with colYA:
    do_preview_yaml = st.button("🧪 Prévisualiser (YAML)")
with colYB:
    do_integrate_yaml = st.button("🧩 Intégrer avec YAML → template")

if do_preview_yaml or do_integrate_yaml:
    try:
        if yaml is None:
            st.error("Le module 'pyyaml' n'est pas installé. Fais: pip install pyyaml")
            st.stop()
        cfg = yaml.safe_load(yaml_text)
        if not cfg:
            st.error("YAML vide ou invalide.")
            st.stop()

        raw, orders = _load_raw_with_colorders()
        # Filtrer par Code DI non vide si présent
        try:
            c_code = get_col(raw, "Custom Attribute5 Value")
            raw[c_code] = raw[c_code].astype(str).str.strip()
            data = raw[raw[c_code].notna() & (raw[c_code] != "")].copy()
        except Exception:
            data = raw.copy()

        # 1) Variables
        var_specs: dict = cfg.get("variables", {}) or {}
        rows_out = []
        for _, row in data.iterrows():
            env: dict[str, float | None] = {}
            # Remplir env de toutes les variables du YAML
            for vname, vspec in var_specs.items():
                env[vname] = _parse_number(_value_from_source_spec(row, orders, vspec))

            # 2) Colonnes directes (pour aperçu on montre juste ce qu'on écrira)
            direct_vals = []
            for item in cfg.get("columns", []) or []:
                target = item.get("target")
                tocc = int(item.get("target_occurrence", 1))
                sval = _value_from_source_spec(row, orders, item)
                direct_vals.append((target, tocc, sval))

            # 3) Calculés
            computed_vals = []
            for item in cfg.get("computed", []) or []:
                target = item.get("target")
                tocc = int(item.get("target_occurrence", 1))
                expr = item.get("expr", "")
                val = _eval_expr(expr, env)
                computed_vals.append((target, tocc, val))

            rows_out.append({
                "__source_file__": row.get("__source_file__"),
                "Code DI": row.get(c_code) if 'c_code' in locals() else None,
                "vars": env,
                "direct": direct_vals,
                "computed": computed_vals,
            })

        # Aperçu convivial: on aplatit un peu
        prev_rows = []
        for r in rows_out[:200]:  # limiter aperçu
            base = {"__source_file__": r["__source_file__"], "Code DI": r.get("Code DI")}
            # Ajouter variables
            for k, v in r["vars"].items():
                base[f"var:{k}"] = v
            # Ajouter direct/computed (concat label#occ)
            for (t, occ, v) in r["direct"]:
                base[f"direct:{t}#{occ}"] = v
            for (t, occ, v) in r["computed"]:
                base[f"calc:{t}#{occ}"] = v
            prev_rows.append(base)
        prev_df = pd.DataFrame(prev_rows)
        st.success(f"Aperçu YAML: {len(prev_df)} lignes (affichage tronqué à 200)")
        st.dataframe(prev_df)

        if do_integrate_yaml:
            # Intégration réelle dans le template
            template_path = Path(TEMPLATE_DEFAULT)
            if 'sheet' in cfg and cfg['sheet']:
                # laisser openpyxl choisir, on sélectionnera la feuille ensuite
                pass
            out_xlsm = copy_template_to_dest(template_path, dest_dir, file_tag, mode)
            wb = load_workbook(out_xlsm, keep_vba=True, data_only=False)

            # Feuille + header row
            ws = wb[cfg.get("sheet")] if cfg.get("sheet") in wb.sheetnames else wb.active
            header_row = int(cfg.get("header_row", 6))
            targets_index = _build_targets_index(ws, header_row)

            start_row = header_row + 1
            rwrite = 0
            for _, row in data.iterrows():
                r = start_row + rwrite
                # (a) Colonnes directes
                for item in cfg.get("columns", []) or []:
                    # cible par header+occurrence ou par target_letter
                    cidx = None
                    if item.get("target_letter"):
                        cidx = column_index_from_string(item["target_letter"])
                    elif item.get("target"):
                        lab = _norm(item["target"])
                        occ = int(item.get("target_occurrence", 1))
                        col_list = targets_index.get(lab, [])
                        if 1 <= occ <= len(col_list):
                            cidx = col_list[occ-1]
                    if cidx is None:
                        continue
                    sval = _value_from_source_spec(row, orders, item)
                    ws.cell(row=r, column=cidx).value = sval

                # (b) Variables env
                env: dict[str, float | None] = {}
                for vname, vspec in var_specs.items():
                    env[vname] = _parse_number(_value_from_source_spec(row, orders, vspec))

                # (c) Computed
                for item in cfg.get("computed", []) or []:
                    cidx = None
                    if item.get("target_letter"):
                        cidx = column_index_from_string(item["target_letter"])
                    elif item.get("target"):
                        lab = _norm(item["target"])
                        occ = int(item.get("target_occurrence", 1))
                        col_list = targets_index.get(lab, [])
                        if 1 <= occ <= len(col_list):
                            cidx = col_list[occ-1]
                    if cidx is None:
                        continue
                    val = _eval_expr(item.get("expr", ""), env)
                    ws.cell(row=r, column=cidx).value = val

                rwrite += 1

            wb.save(out_xlsm)
            st.success(f"✅ Intégration YAML terminée → {out_xlsm}")
            with open(out_xlsm, "rb") as f:
                st.download_button(
                    label="⬇️ Télécharger le template rempli (.xlsm)",
                    data=f.read(),
                    file_name=out_xlsm.name,
                    mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                )

    except Exception as e:
        st.exception(e)
