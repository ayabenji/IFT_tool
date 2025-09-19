from __future__ import annotations

import re
from datetime import date
from pathlib import Path
from typing import Iterable
from dataclasses import dataclass

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet


from excel_read import _norm


DATE_IN_NAME = re.compile(r"(\d{2})-(\d{2})-(\d{4})")
IFT_FILE_PATTERNS = ("IFT - *.xlsm", "IFT - *.xlsx")
ANALYSIS_SHEET = "IRS - INF – XCCY"
ANALYSIS_HEADER_ROW = 6
ANALYSIS_DIRTY_LETTER = "AN"

def _norm_aliases(*labels: str) -> set[str]:
    return {_norm(label) for label in labels}


@dataclass(frozen=True)
class MtMSummaryItem:
    label: str
    categories: tuple[str, ...]
    bullet: str
    needs_reason: bool
    fallback_aliases: tuple[str, ...] = ()


CATEGORY_ALIASES: dict[str, set[str]] = {
    "IR_SWAP": _norm_aliases("IR Swap", "IR - Swap", "CMS Swap"),
    "ASWI": _norm_aliases("ASWI", "IR ASWI", "Real Rate Swap"),
    "XCCY": _norm_aliases("XCCY", "IR XCCY", "Cross Currency Swap"),
    "BOND_FWD": _norm_aliases("Bond Forward", "Bond Fwd", "Bond FWD", "BondForward"),
}


MTM_SUMMARY_ITEMS: tuple[MtMSummaryItem, ...] = (
    MtMSummaryItem(
        label="IR – ASWI – XCCY",
        categories=("IR_SWAP", "ASWI", "XCCY"),
        bullet="-  ",
        needs_reason=False,
        fallback_aliases=tuple(sorted(_norm_aliases("IR – ASWI – XCCY", "IR - ASWI - XCCY", "IR-ASWI-XCCY")))),
    MtMSummaryItem(
        label="IR Swap",
        categories=("IR_SWAP",),
        bullet="- ",
        needs_reason=True,
    ),
    MtMSummaryItem(
        label="ASWI",
        categories=("ASWI",),
        bullet="- ",
        needs_reason=True,
    ),
        MtMSummaryItem(
        label="XCCY",
        categories=("XCCY",),
        bullet="o ",
        needs_reason=True,
    ),
        MtMSummaryItem(
        label="Bond Forward",
        categories=("BOND_FWD",),
        bullet="- ",
        needs_reason=True,
    ),
)

    

def _extract_date_from_name(name: str) -> date | None:
    match = DATE_IN_NAME.search(name)
    if not match:
        return None
    day, month, year = map(int, match.groups())
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _previous_quarter_folder(ifts_dt: date) -> tuple[int, int]:
    quarter = (ifts_dt.month - 1) // 3 + 1
    prev_quarter = quarter - 1
    if prev_quarter == 0:
        prev_quarter = 4
        year = ifts_dt.year - 1
    else:
        year = ifts_dt.year
    quarter_end_months = (3, 6, 9, 12)
    month = quarter_end_months[prev_quarter - 1]
    return year, month


def _iter_existing_files(dirs: Iterable[Path]) -> list[Path]:
    results: list[Path] = []
    for folder in dirs:
        if not folder or not folder.exists():
            continue
        for pattern in IFT_FILE_PATTERNS:
            results.extend(sorted(folder.glob(pattern)))
    return results


def _pick_latest_file(paths: Iterable[Path]) -> Path | None:
    dated: list[tuple[date, Path]] = []
    undated: list[Path] = []
    for path in paths:
        extracted = _extract_date_from_name(path.name)
        if extracted:
            dated.append((extracted, path))
        else:
            undated.append(path)
    if dated:
        dated.sort(key=lambda x: (x[0], x[1].name))
        return dated[-1][1]
    if undated:
        undated.sort()
        return undated[-1]
    return None


def locate_previous_production(base_out: Path, mode: str, ifts_dt: date) -> Path | None:
    year_folder = ifts_dt.strftime("%Y")
    month_folder = ifts_dt.strftime("%m-%Y")
    candidates: list[Path] = []
    if mode.lower() == "close":
        month_dir = base_out / year_folder / month_folder
        candidates.extend(
            [
                month_dir / "prod" / "fast",
                month_dir / "fast",
                month_dir,
            ]
        )
    else:
        prev_year, prev_month = _previous_quarter_folder(ifts_dt)
        prev_month_folder = f"{prev_month:02d}-{prev_year}"
        quarter_dir = base_out / str(prev_year) / prev_month_folder
        candidates.extend(
            [
                quarter_dir / "prod" / "close",
                quarter_dir / "close",
                quarter_dir,
            ]
        )
    paths = _iter_existing_files(candidates)
    return _pick_latest_file(paths)


def guess_current_production(dest_dir: Path, file_tag: str, mode: str) -> Path | None:
    if not dest_dir.exists():
        return None
    base_name = f"IFT_{file_tag}_{mode.lower()}"
    for ext in (".xlsm", ".xlsx"):
        candidate = dest_dir / f"{base_name}{ext}"
        if candidate.exists():
            return candidate
    candidates = sorted(dest_dir.glob("IFT*.xls*"))
    if candidates:
        return candidates[-1]
    return None


def _to_float(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace("\xa0", "").replace(" ", "").strip()
        if not cleaned:
            return None
        if cleaned.count(",") == 1 and cleaned.count(".") == 0:
            cleaned = cleaned.replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def aggregate_dirty_by_classif(path: Path) -> dict[str, float]:
    wb = load_workbook(path, data_only=True, read_only=True)
    bond_forward_total: float | None = None
    totals: dict[str, float] = {}
    try:
        if ANALYSIS_SHEET not in wb.sheetnames:
            raise KeyError(f"Feuille '{ANALYSIS_SHEET}' absente de {path.name}")
        ws = wb[ANALYSIS_SHEET]
        classif_col = None
        max_col = ws.max_column
        for col in range(1, max_col + 1):
            header_val = ws.cell(row=ANALYSIS_HEADER_ROW, column=col).value
            if header_val is None:
                continue
            if _norm(header_val) == _norm("Classif DI"):
                classif_col = col
                break
        if classif_col is None:
            raise KeyError("Colonne 'Classif DI' introuvable")
        dirty_idx = column_index_from_string(ANALYSIS_DIRTY_LETTER)
        for row in ws.iter_rows(
            min_row=ANALYSIS_HEADER_ROW + 1,
            max_row=ws.max_row,
            values_only=True,
        ):
            if classif_col - 1 >= len(row) or dirty_idx - 1 >= len(row):
                continue
            classif_raw = row[classif_col - 1]
            dirty_raw = row[dirty_idx - 1]
            if classif_raw is None:
                continue
            classif = str(classif_raw).strip()
            if not classif:
                continue
            dirty = _to_float(dirty_raw)
            if dirty is None:
                continue
            totals[classif] = totals.get(classif, 0.0) + dirty
        if "BND FWD" in wb.sheetnames:
            bond_forward_total = _sum_bnd_fwd_prix_gam(wb["BND FWD"])
    finally:
        wb.close()
    if bond_forward_total is not None:
        totals["Bond Forward"] = totals.get("Bond Forward", 0.0) + bond_forward_total
    return totals

def _find_column_by_header(
    ws: Worksheet, header: str, *, max_search_rows: int = 12
) -> tuple[int | None, int | None]:
    target = _norm(header)
    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=1,
            max_row=min(max_search_rows, ws.max_row),
            max_col=ws.max_column,
            values_only=True,
        ),
        start=1,
    ):
        for col_idx, value in enumerate(row, start=1):
            if value is None:
                continue
            if _norm(str(value)) == target:
                return col_idx, row_idx
    return None, None


def _sum_bnd_fwd_prix_gam(ws: Worksheet) -> float | None:
    prix_col, header_row = _find_column_by_header(ws, "Prix GAM")
    if prix_col is None or header_row is None:
        return None
    total = 0.0
    found = False
    for row in ws.iter_rows( min_row=header_row + 1,
        max_row=ws.max_row,
        values_only=True,
    ):
        if prix_col - 1 >= len(row):
            continue
        value = row[prix_col - 1]
        numeric = _to_float(value)
        if numeric is None:
            continue
        total += numeric
        found = True
    return total if found else None


    

    
    

def build_comparison_dataframe(
    current_map: dict[str, float], previous_map: dict[str, float]
) -> pd.DataFrame:
    keys = sorted({*current_map.keys(), *previous_map.keys()})
    rows = []
    for key in keys:
        old_val = previous_map.get(key, 0.0)
        new_val = current_map.get(key, 0.0)
        delta = new_val - old_val
        pct = (delta / old_val * 100.0) if abs(old_val) > 1e-12 else None
        rows.append(
            {
                "Classif DI": key,
                "Dirty ancienne prod": old_val,
                "Dirty prod actuelle": new_val,
                "Écart": delta,
                "Écart %": pct,
            }
        )
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values("Écart", key=lambda s: s.abs(), ascending=False)
    return df

def _prepare_norm_totals(totals: dict[str, float]) -> dict[str, float]:
    return {_norm(key): value for key, value in totals.items()}


def _lookup_alias_total(
    norm_totals: dict[str, float], aliases: Iterable[str]
) -> float | None:
    for alias in aliases:
        if alias in norm_totals:
            return norm_totals[alias]
    return None

def _resolve_summary_value(
    norm_totals: dict[str, float], item: MtMSummaryItem
) -> float | None:
    total = 0.0
    found = False
    for category in item.categories:
        aliases = CATEGORY_ALIASES.get(category)
        if not aliases:
            continue
        category_value = _lookup_alias_total(norm_totals, aliases)
        if category_value is None:
            continue
        total += category_value
        found = True
    if found:
        return total
    if item.fallback_aliases:
        return _lookup_alias_total(norm_totals, item.fallback_aliases)
    return None

def _format_mtm_value(value: float | None) -> str:
    if value is None:
        return "N/A"
    scaled = value
    if abs(value) >= 1e4:
        scaled = value / 1_000_000
    sign = "- " if scaled < 0 else ""
    body = f"{abs(scaled):,.1f}".replace(",", " ")
    return f"{sign}{body}" if sign else body


def build_mtm_summary(
    current_totals: dict[str, float],
    previous_totals: dict[str, float],
    current_date: date | None,
    previous_date: date | None,
) -> str | None:
    norm_current = _prepare_norm_totals(current_totals)
    norm_previous = _prepare_norm_totals(previous_totals)
    lines: list[str] = ["Bonjour,", ""]
    has_content = False
    date_str = current_date.strftime("%d/%m/%Y") if current_date else ""
    prev_date_str = previous_date.strftime("%d/%m/%Y") if previous_date else ""
    
    lines.append(f"Voici le fichier des IFTs au {date_str}  :\n")
    lines.append("I. Point MtM :\n")
    for item in MTM_SUMMARY_ITEMS:
        cur_val = _resolve_summary_value(norm_current, item)
        prev_val = _resolve_summary_value(norm_previous, item)
        if cur_val is None and prev_val is None:
            continue
        has_content = True
        cur_txt = _format_mtm_value(cur_val)
        prev_txt = _format_mtm_value(prev_val)
        suffix = ""
        if date_str:
            suffix = f" MEUR au {prev_date_str} \n"
        else:
            suffix = " MEUR"

        lines.append(f"{item.bullet}{item.label} : {cur_txt} vs {prev_txt}{suffix}")
        if item.needs_reason:
            lines.append("-> Dégradation/Amélioration du MTM en raison (à remplir)")
        lines.append("")
    if not has_content:
        return None
    return "\n".join(lines).strip()

def render_analysis_tab(
    base_out: Path,
    dest_dir: Path,
    mode: str,
    ifts_date: date,
    file_tag: str,
) -> None:
    st.write(
        "Comparer la production actuelle avec la précédente pour identifier les écarts de dirty value par `Classif DI`."
    )

    previous_guess = locate_previous_production(base_out, mode, ifts_date)
    current_guess = guess_current_production(dest_dir, file_tag, mode)
    current_session_path = Path(st.session_state.get("out_xlsm", "")) if "out_xlsm" in st.session_state else None
    if current_session_path and current_session_path.exists():
        current_guess = current_session_path

    current_path_str = st.text_input(
        "Fichier de prod actuel (.xlsm)",
        value=str(current_guess) if current_guess else "",
        placeholder="Chemin complet du fichier courant",
    )
    previous_path_str = st.text_input(
        "Ancienne production à comparer",
        value=str(previous_guess) if previous_guess else "",
        placeholder="Chemin complet du fast/close précédent",
    )

    if st.button("🔍 Lancer l'analyse comparative"):
        try:
            if not current_path_str:
                st.error("Indique le fichier de production actuel à analyser.")
                st.stop()
            if not previous_path_str:
                st.error("Indique le fichier de l'ancienne production (fast/close).")
                st.stop()
            current_path = Path(current_path_str)
            previous_path = Path(previous_path_str)
            if not current_path.exists():
                st.error(f"Fichier actuel introuvable : {current_path}")
                st.stop()
            if not previous_path.exists():
                st.error(f"Ancienne production introuvable : {previous_path}")
                st.stop()

            st.write(f"Fichier actuel : **{current_path.name}**")
            st.write(f"Ancienne prod : **{previous_path.name}**")

            current_totals = aggregate_dirty_by_classif(current_path)
            previous_totals = aggregate_dirty_by_classif(previous_path)

            df_compare = build_comparison_dataframe(current_totals, previous_totals)
            if df_compare.empty:
                st.warning("Aucune donnée agrégée — vérifier la feuille ou les colonnes cibles.")
                st.stop()

            total_current = sum(current_totals.values())
            total_previous = sum(previous_totals.values())
            delta_total = total_current - total_previous
            col_cur, col_prev = st.columns(2)
            with col_cur:
                st.metric(
                    "Total dirty actuel",
                    f"{total_current:,.2f}",
                    delta=f"{delta_total:,.2f}",
                )
            with col_prev:
                st.metric("Total dirty ancien", f"{total_previous:,.2f}")

            st.dataframe(df_compare)


            current_file_date = _extract_date_from_name(current_path.name)
            previous_file_date = _extract_date_from_name(previous_path.name)
            summary_text = build_mtm_summary(
                current_totals,
                previous_totals,
                current_file_date or ifts_date,
                previous_file_date or current_file_date or ifts_date,
            )
            if summary_text:
                st.markdown(summary_text)

        except Exception as exc:
            st.exception(exc)