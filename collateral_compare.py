from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Mapping

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import column_index_from_string


def _to_float(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace("\xa0", "").replace(" ", "").strip()
        if not cleaned:
            return None
        if cleaned.count(",") == 1 and cleaned.count(".") == 0:
            cleaned = cleaned.replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _clean_label(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
    else:
        text = str(value).strip()
    if not text:
        return None
    return " ".join(text.split())


def parse_alias_mapping(text: str) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if "#" in line:
            line = line.split("#", 1)[0].strip()
            if not line:
                continue

        alias: str | None = None
        canonical: str | None = None
        if "=" in line:
            alias, canonical = line.split("=", 1)
        elif ":" in line:
            alias, canonical = line.split(":", 1)
        elif "\t" in line:
            parts = [part.strip() for part in line.split("\t") if part.strip()]
            if len(parts) >= 2:
                alias, canonical = parts[0], parts[1]
        if alias is None or canonical is None:
            continue

        alias_clean = _clean_label(alias)
        canonical_clean = _clean_label(canonical)
        if not alias_clean or not canonical_clean:
            continue

        mapping[alias_clean.casefold()] = canonical_clean
        mapping.setdefault(canonical_clean.casefold(), canonical_clean)
    return mapping


def normalize_label(value, mapping: Mapping[str, str] | None = None) -> str | None:
    cleaned = _clean_label(value)
    if cleaned is None:
        return None
    if mapping:
        canonical = mapping.get(cleaned.casefold())
        if canonical:
            return canonical
    return cleaned.upper()

def _first_notna(series: pd.Series):
    for value in series:
        if pd.notna(value):
            return value
    return None


def _sum_notna(series: pd.Series) -> float:
    total = 0.0
    has_value = False
    for value in series:
        if pd.isna(value):
            continue
        total += float(value)
        has_value = True
    return total if has_value else 0.0


def _aggregate_by_norm(
    df: pd.DataFrame,
    *,
    text_columns: Iterable[str],
    value_columns: Iterable[str],
) -> pd.DataFrame:
    if df.empty:
        return df
    agg_spec: dict[str, object] = {}
    for col in text_columns:
        if col in df.columns:
            agg_spec[col] = _first_notna
    for col in value_columns:
        if col in df.columns:
            agg_spec[col] = _sum_notna
    if not agg_spec:
        return df
    grouped = (
        df.groupby(["Norm Counterparty", "Norm Typologie"], as_index=False).agg(
            agg_spec
        )
    )
    return grouped

def find_collateral_report(dest_dir: Path) -> Path:
    candidates = sorted(dest_dir.glob("*Report Collatéral.xlsx"))
    if not candidates:
        raise FileNotFoundError(
            f"Aucun fichier '*Report Collatéral.xlsx' trouvé dans {dest_dir}"
        )
    return max(candidates, key=lambda p: p.stat().st_mtime)


def load_collateral_summary(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Report Collateral", dtype=object)

    expected = ["Counterparty ", "Typologie ", "MTM Gam ", "MTM Counterparty "]
    rename= ["Counterparty", "Typologie", "MtM Gam", "MtM Counterparty"]
    missing = [col for col in expected if col not in df.columns]
    if missing:
        raise KeyError(
            f"Colonnes manquantes dans {path.name} (Report Collateral) : {', '.join(missing)}"
        )
    df = df[expected].copy()
    df.rename(columns=dict(zip(expected, rename)), inplace=True)
    for col in ["Counterparty", "Typologie"]:
        df[col] = df[col].map(_clean_label)
    for col in ["MtM Gam", "MtM Counterparty"]:
        df[col] = df[col].map(_to_float)
    df = df.dropna(subset=["Counterparty", "Typologie"], how="any")
    return df


@dataclass
class _AggregationSpec:
    sheet_name: str
    classif_letter: str | None
    counterparty_letter: str
    mtm_gam_letter: str
    mtm_counterparty_letter: str
    classif_override: str | None = None


def aggregate_template_mtm(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        specs: Iterable[_AggregationSpec] = (
            _AggregationSpec(
                sheet_name="IRS - INF – XCCY",
                classif_letter="B",
                counterparty_letter="E",
                mtm_gam_letter="AN",
                mtm_counterparty_letter="AW",
            ),
            _AggregationSpec(
                sheet_name="BND FWD",
                classif_letter=None,
                counterparty_letter="C",
                mtm_gam_letter="E",
                mtm_counterparty_letter="F",
                classif_override="Forward",
            ),
        )
        aggregates: dict[tuple[str, str], dict[str, float]] = {}
        for spec in specs:
            if spec.sheet_name not in wb.sheetnames:
                continue
            ws = wb[spec.sheet_name]
            _aggregate_sheet(ws, spec, aggregates)
    finally:
        wb.close()

    rows: list[dict[str, object]] = []
    for (classif, counterparty), data in sorted(aggregates.items()):
        rows.append(
            {
                "Classif DI": classif,
                "Counterparty": counterparty,
                "MtM Gam": data.get("MtM Gam", 0.0),
                "MtM Counterparty": data.get("MtM Counterparty", 0.0),
            }
        )
    return pd.DataFrame(rows)


def _aggregate_sheet(
    ws: Worksheet,
    spec: _AggregationSpec,
    aggregates: dict[tuple[str, str], dict[str, float]],
) -> None:
    classif_idx = (
        column_index_from_string(spec.classif_letter)
        if spec.classif_letter
        else None
    )
    counterparty_idx = column_index_from_string(spec.counterparty_letter)
    gam_idx = column_index_from_string(spec.mtm_gam_letter)
    cp_idx = column_index_from_string(spec.mtm_counterparty_letter)

    start_row = 2
    if classif_idx is not None:
        header_row = _find_header_row(ws, classif_idx, "Classif DI")
        start_row = max(start_row, header_row + 1)
    else:
        header_row = _find_header_row(ws, counterparty_idx, "Counterparty")
        if header_row:
            start_row = max(start_row, header_row + 1)

    for values in ws.iter_rows(
        min_row=start_row,
        max_row=ws.max_row,
        values_only=True,
    ):
        row_len = len(values)
        if classif_idx is not None and classif_idx - 1 >= row_len:
            continue
        if counterparty_idx - 1 >= row_len:
            continue
        if gam_idx - 1 >= row_len and cp_idx - 1 >= row_len:
            continue

        classif_raw = spec.classif_override
        if classif_raw is None and classif_idx is not None:
            classif_raw = values[classif_idx - 1]
        counterparty_raw = values[counterparty_idx - 1]
        classif = _clean_label(classif_raw)
        counterparty = _clean_label(counterparty_raw)
        if classif is None or counterparty is None:
            continue
        if counterparty.casefold() in {"counterparty", "contrepartie", "cp", "total"}:
            continue

        mtm_gam = None
        if gam_idx - 1 < row_len:
            mtm_gam = _to_float(values[gam_idx - 1])
        mtm_counterparty = None
        if cp_idx - 1 < row_len:
            mtm_counterparty = _to_float(values[cp_idx - 1])

        if mtm_gam is None and mtm_counterparty is None:
            continue

        key = (classif, counterparty)
        slot = aggregates.setdefault(key, {"MtM Gam": 0.0, "MtM Counterparty": 0.0})
        if mtm_gam is not None:
            slot["MtM Gam"] += mtm_gam
        if mtm_counterparty is not None:
            slot["MtM Counterparty"] += mtm_counterparty


def _find_header_row(ws: Worksheet, column_index: int, expected: str) -> int:
    expected_norm = expected.casefold()
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row=row, column=column_index).value
        if value is None:
            continue
        cleaned = _clean_label(value)
        if cleaned and cleaned.casefold() == expected_norm:
            return row
    return 1


def build_collateral_comparison(
    template_df: pd.DataFrame,
    collateral_df: pd.DataFrame,
    *,
    counterparty_aliases: Mapping[str, str] | None = None,
    typology_aliases: Mapping[str, str] | None = None,
) -> pd.DataFrame:
    if template_df.empty and collateral_df.empty:
        return pd.DataFrame()

    tmpl = template_df.copy()
    coll = collateral_df.copy()

    tmpl["Norm Counterparty"] = tmpl["Counterparty"].map(
        lambda v: normalize_label(v, counterparty_aliases)
    )
    tmpl["Norm Typologie"] = tmpl["Classif DI"].map(
        lambda v: normalize_label(v, typology_aliases)
    )
    coll["Norm Counterparty"] = coll["Counterparty"].map(
        lambda v: normalize_label(v, counterparty_aliases)
    )
    coll["Norm Typologie"] = coll["Typologie"].map(
        lambda v: normalize_label(v, typology_aliases)
    )

    tmpl = tmpl.dropna(subset=["Norm Counterparty", "Norm Typologie"], how="any")
    coll = coll.dropna(subset=["Norm Counterparty", "Norm Typologie"], how="any")

    tmpl = _aggregate_by_norm(
        tmpl,
        text_columns=("Counterparty", "Classif DI"),
        value_columns=("MtM Gam", "MtM Counterparty"),
    )
    coll = _aggregate_by_norm(
        coll,
        text_columns=("Counterparty", "Typologie"),
        value_columns=("MtM Gam", "MtM Counterparty"),
    )

    merged = coll.merge(
        tmpl,
        how="outer",
        on=["Norm Counterparty", "Norm Typologie"],
        suffixes=("_collateral", "_template"),
    )

    merged["MtM Gam_collateral"].fillna(0.0, inplace=True)
    merged["MtM Counterparty_collateral"].fillna(0.0, inplace=True)
    merged["MtM Gam_template"].fillna(0.0, inplace=True)
    merged["MtM Counterparty_template"].fillna(0.0, inplace=True)

    merged["Ecart MtM Gam"] = (
        merged["MtM Gam_template"] - merged["MtM Gam_collateral"]
    )
    merged["Ecart MtM Counterparty"] = (
        merged["MtM Counterparty_template"]
        - merged["MtM Counterparty_collateral"]
    )

    merged["Présent template"] = merged["Counterparty_template"].notna()
    merged["Présent collat"] = merged["Counterparty_collateral"].notna()

    ordered_cols = [
        "Norm Counterparty",
        "Norm Typologie",
        "Counterparty_template",
        "Classif DI",
        "MtM Gam_template",
        "MtM Counterparty_template",
        "Counterparty_collateral",
        "Typologie",
        "MtM Gam_collateral",
        "MtM Counterparty_collateral",
        "Ecart MtM Gam",
        "Ecart MtM Counterparty",
        "Présent template",
        "Présent collat",
    ]

    for col in ordered_cols:
        if col not in merged.columns:
            merged[col] = None

    result = merged[ordered_cols].copy()
    return result.sort_values(
        by=["Norm Counterparty", "Norm Typologie"],
        kind="mergesort",
        na_position="last",
    )