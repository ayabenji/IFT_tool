from __future__ import annotations
from pathlib import Path
from datetime import date

try:
    import win32com.client as win32  # type: ignore
    import pythoncom  # type: ignore
except Exception:  # pragma: no cover
    win32 = None
    pythoncom = None

from openpyxl import load_workbook

__all__ = ["build_ifts_filename", "export_xlsx_copy", "prepare_outlook_draft", "prepare_trioptima_request_mail", "prepare_collateral_report_request_mail"]


def build_ifts_filename(ifts_date: date) -> str:
    """Nom final de la PJ .xlsx, sans macros."""
    return f"VALO IFT {ifts_date.strftime('%d-%m-%Y')} - contrôle SD vs SCD.xlsx"


def export_xlsx_copy(xlsm_path: Path, final_name: str) -> Path:
    """Produit une PJ .xlsx propre à partir du .xlsm rempli.
    1) Essaye via Excel COM (SaveAs .xlsx) → macros retirées proprement, formules recalculées.
    2) Fallback openpyxl: recopie valeurs + formats (sans macros ni contrôles).
    """
    out_path = xlsm_path.parent / final_name

    # 0) Si le fichier source est déjà un .xlsx → simple copie robuste (pas de COM)
    if xlsm_path.suffix.lower() == ".xlsx":
        import shutil, time, os
        for _ in range(5):
            try:
                shutil.copy2(xlsm_path, out_path)
                break
            except PermissionError:
                time.sleep(0.4)
        else:
            # Dernier recours: tenter une copie via un fichier temporaire
            tmp_copy = xlsm_path.with_suffix(".tmp_copy.xlsx")
            try:
                shutil.copy2(xlsm_path, tmp_copy)
                shutil.copy2(tmp_copy, out_path)
                os.remove(tmp_copy)
            except Exception:
                pass
        return out_path


    # 1) Excel COM
    if win32 is not None and pythoncom is not None:
        try:
            pythoncom.CoInitialize()
            excel = win32.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(str(xlsm_path))
            try:
                wb.Application.CalculateFullRebuild()
            except Exception:
                pass
            xlOpenXMLWorkbook = 51  # .xlsx
            wb.SaveAs(str(out_path), FileFormat=xlOpenXMLWorkbook)
            wb.Close(SaveChanges=False)
            excel.Quit()
            return out_path
        except Exception:
            try:
                excel.Quit()
            except Exception:
                pass
        finally:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    # 2) Fallback openpyxl (valeurs + formats)
    from openpyxl import Workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils.cell import column_index_from_string

    src = load_workbook(xlsm_path, data_only=True, keep_vba=False)
    dst = Workbook()
    default_ws = dst.active
    dst.remove(default_ws)

    for ws in src.worksheets:
        new_ws = dst.create_sheet(title=ws.title)
        # Largeurs de colonnes
        for key, dim in ws.column_dimensions.items():
            new_ws.column_dimensions[key].width = dim.width
        # Fusions
        try:
            for mr in ws.merged_cells.ranges:
                new_ws.merge_cells(str(mr))
        except Exception:
            pass
        # Valeurs + number_format
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c, MergedCell):
                    continue
                col_index = getattr(c, "col_idx", None)
                if col_index is None:
                    col_index = c.column if isinstance(c.column, int) else column_index_from_string(c.column)
                nc = new_ws.cell(row=c.row, column=col_index, value=c.value)
                try:
                    nc.number_format = c.number_format
                except Exception:
                    pass
        # Gel de volets
        try:
            new_ws.freeze_panes = ws.freeze_panes
        except Exception:
            pass

    dst.save(out_path)
    return out_path


def prepare_outlook_draft(attachment_path: Path, ifts_date: date, to: str = "", cc: str = "") -> None:
    """Crée un brouillon Outlook avec PJ, objet et corps préremplis."""
    if win32 is None or pythoncom is None:
        raise RuntimeError("Le module 'pywin32' n'est pas installé. Fais: pip install pywin32")
    pythoncom.CoInitialize()
    try:
        # Utiliser l'instance existante d'Outlook si possible (évite les blocages)
        try:
            outlook = win32.gencache.EnsureDispatch("Outlook.Application")
        except Exception:
            outlook = win32.DispatchEx("Outlook.Application")
        mail = outlook.CreateItem(0)  # olMailItem
        date_str = ifts_date.strftime('%d/%m/%Y')
        mail.Subject = f"VALO IFT – Contrôle SD vs SCD – {date_str}"
        mail.Body = (
            f"Bonjour Sebastien,\n\n"
            f"Voici nos données SD au {date_str}. Pourrais-tu vérifier avec celles qui alimentent Dimension ?\n\n"
            "Merci,"
        )
        if to:
            mail.To = to
        if cc:
            mail.CC = cc
        mail.Attachments.Add(str(attachment_path))
        # Affichage **modèle non bloquant** pour ne pas figer Outlook
        try:
            mail.Display(False)
        except TypeError:
            # Certains environnements n'acceptent pas l'argument → appel sans paramètre
            mail.Display()
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass

def prepare_trioptima_request_mail(ifts_date: date, to: str = "", cc: str = "") -> None:
    """Crée un brouillon Outlook pour demander le fichier Trioptima du jour IFTs."""
    if win32 is None or pythoncom is None:
        raise RuntimeError("Le module 'pywin32' n'est pas installé. Fais: pip install pywin32")
    pythoncom.CoInitialize()
    try:
        try:
            outlook = win32.gencache.EnsureDispatch("Outlook.Application")
        except Exception:
            outlook = win32.DispatchEx("Outlook.Application")
        mail = outlook.CreateItem(0)
        date_str = ifts_date.strftime('%d/%m/%Y')
        mail.Subject = f"IFT {date_str} - Fichier trioptima - search_groupama-am.csv"
        mail.Body = (
            "Bonjour,\n\n"
            "Dans le cadre de l’exercice de la contre-valorisation trimestrielle des Instruments financiers à Terme"
            f"Pourriez-vous m’envoyer, svp, le fichier Trioptima  (search_groupama-am.csv) du {date_str} dès qu’il sera disponible ? \n\n"
            "Merci d’avance."
        )
        if to:
            mail.To = to
        if cc:
            mail.CC = cc
        try:
            mail.Display(False)
        except TypeError:
            mail.Display()
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass

def prepare_collateral_report_request_mail(ifts_date: date, to: str = "", cc: str = "") -> None:
    """Crée un brouillon Outlook pour demander le report trimestriel de collatéral."""
    if win32 is None or pythoncom is None:
        raise RuntimeError("Le module 'pywin32' n'est pas installé. Fais: pip install pywin32")
    pythoncom.CoInitialize()
    try:
        try:
            outlook = win32.gencache.EnsureDispatch("Outlook.Application")
        except Exception:
            outlook = win32.DispatchEx("Outlook.Application")
        mail = outlook.CreateItem(0)
        date_str = ifts_date.strftime('%d/%m/%Y')
        mail.Subject = f"Report trimestriel de collatéral {date_str}"
        mail.Body = (
            "Bonjour,\n\n"
            f"Dans le cadre des IFT, pouvez-vous nous envoyer le report trimestriel de collatéral au {date_str}? \n\n "
            "Merci"
        )
        if to:
            mail.To = to
        if cc:
            mail.CC = cc
        try:
            mail.Display(False)
        except TypeError:
            mail.Display()
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass