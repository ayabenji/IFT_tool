from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Dict, Iterable, Tuple

from openpyxl import load_workbook

from excel_read import _norm
from template_write import build_targets_index, letter_to_index

__all__ = [
    "expected_sensis_name",
    "locate_sensis_file",
    "load_sensis_table",
    "apply_sensis_to_workbook",
]


@dataclass
class SensisEntry:
    dirty_pct: float | None
    clean_pct: float | None
    accrued_pct: float | None
    sensis_leg1: float | None
    sensis_leg2: float | None
    duration_leg1: float | None
    duration_leg2: float | None
    duration_total: float | None

    @property
    def sensis_total(self) -> float | None:
        vals = [v for v in (self.sensis_leg1, self.sensis_leg2) if v is not None]
        if not vals:
            return None
        return sum(vals)


def expected_sensis_name(ifts_date: date, suffix: str = "xlsx") -> str:
    return f"Sensis IFTTool_{ifts_date:%d%m%Y}.{suffix.lower()}"


def locate_sensis_file(dest_dir: Path, ifts_date: date) -> Path:
    candidates = [
        dest_dir / expected_sensis_name(ifts_date, "xlsx"),
        dest_dir / expected_sensis_name(ifts_date, "xlsm"),
    ]
    for path in candidates:
        if path.exists():
            return path
    pattern = f"Sensis IFTTool_{ifts_date:%d%m%Y}*.xls*"
    matches = sorted(dest_dir.glob(pattern))
    if matches:
        return matches[0]
    raise FileNotFoundError(f"Fichier Sensis introuvable dans {dest_dir}: {pattern}")


def _normalize_header(value: object) -> str:
    text = " ".join(str(value).strip().lower().split()) if value is not None else ""
    text = (
        text.replace("(", " ")
        .replace(")", " ")
        .replace("/", " ")
        .replace("%", " %")
    )
    text = " ".join(text.split())
    return text


_HEADER_ALIASES: Dict[str, Tuple[str, ...]] = {
    "code": ("code di",),
    "notional": ("notional",),
    "dirty_pct": ("dirty price %", "dirty price(%)"),
    "clean_pct": ("clean price %", "clean price(%)"),
    "accrued_pct": ("couru %", "accrued interest %"),
}


def _find_header_row(ws, search_rows: int = 12) -> int:
    max_row = min(ws.max_row, search_rows)
    max_col = ws.max_column
    for row in range(1, max_row + 1):
        values = [ws.cell(row=row, column=col).value for col in range(1, max_col + 1)]
        normed = {_normalize_header(v) for v in values if v is not None}
        if any(alias in normed for alias in ("code di", "code")):
            return row
    return 3


def _build_header_map(ws, header_row: int) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is None:
            continue
        key = _normalize_header(value)
        if key and key not in mapping:
            mapping[key] = col
    return mapping


def _column_for(header_map: Dict[str, int], logical: str) -> int | None:
    aliases = _HEADER_ALIASES.get(logical, ())
    for alias in aliases:
        if alias in header_map:
            return header_map[alias]
    return None


def _column_from_letter(letter: str) -> int:
    return letter_to_index(letter) + 1


def load_sensis_table(path: Path, sheet: str = "Valorisation - IFT - Valeur") -> Dict[str, SensisEntry]:
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        raise KeyError(f"Feuille '{sheet}' absente de {path.name}")
    ws = wb[sheet]
    header_row = _find_header_row(ws)
    header_map = _build_header_map(ws, header_row)

    col_code = _column_for(header_map, "code")
    if col_code is None:
        raise KeyError("Colonne 'Code DI' introuvable dans Sensis")

    col_dirty = _column_for(header_map, "dirty_pct")
    col_clean = _column_for(header_map, "clean_pct")
    col_accrued = _column_for(header_map, "accrued_pct")

    col_sensis1 = _column_from_letter("AC")
    col_sensis2 = _column_from_letter("AD")
    col_dur1 = _column_from_letter("AE")
    col_dur2 = _column_from_letter("AF")
    col_dur_total = _column_from_letter("Z")

    data: Dict[str, SensisEntry] = {}
    for row in range(header_row + 1, ws.max_row + 1):
        code_cell = ws.cell(row=row, column=col_code).value
        if code_cell is None:
            continue
        code = str(code_cell).strip()
        if not code:
            continue
        entry = SensisEntry(
            dirty_pct=ws.cell(row=row, column=col_dirty).value if col_dirty else None,
            clean_pct=ws.cell(row=row, column=col_clean).value if col_clean else None,
            accrued_pct=ws.cell(row=row, column=col_accrued).value if col_accrued else None,
            sensis_leg1=ws.cell(row=row, column=col_sensis1).value,
            sensis_leg2=ws.cell(row=row, column=col_sensis2).value,
            duration_leg1=ws.cell(row=row, column=col_dur1).value,
            duration_leg2=ws.cell(row=row, column=col_dur2).value,
            duration_total=ws.cell(row=row, column=col_dur_total).value,
        )
        data[code] = entry
    return data


def _to_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\u00a0", " ")
    if not text:
        return None
    text = text.replace(" ", "").replace("'", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _mul(a: object, b: object) -> float | None:
    fa = _to_float(a)
    fb = _to_float(b)
    if fa is None or fb is None:
        return None
    return fa * fb


def _sum_optional(values: Iterable[object]) -> float | None:
    collected = [v for v in (_to_float(x) for x in values) if v is not None]
    if not collected:
        return None
    return sum(collected)


def _sub_optional(left: object, right: object) -> float | None:
    lf = _to_float(left)
    rf = _to_float(right)
    if lf is None and rf is None:
        return None
    return (lf or 0.0) - (rf or 0.0)


def apply_sensis_to_workbook(
    workbook_path: Path,
    sensis: Dict[str, SensisEntry],
    header_row: int = 6,
    sheet: str = "IRS - INF – XCCY",
) -> Tuple[int, list[str], list[dict[str, object]]]:
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    wb = load_workbook(workbook_path, keep_vba=keep_vba, data_only=False)
    if sheet not in wb.sheetnames:
        raise KeyError(f"Feuille '{sheet}' absente de {workbook_path.name}")
    ws = wb[sheet]

    targets = build_targets_index(ws, header_row)
    code_cols = targets.get(_norm("Code DI"), [])
    if not code_cols:
        raise KeyError("Colonne 'Code DI' introuvable dans le template")
    code_col = code_cols[0]

    start_row = header_row + 1
    updated = 0
    missing_codes: list[str] = []
    updated_rows: list[dict[str, object]] = []

    col_notional = letter_to_index("M") + 1

    for row in range(start_row, ws.max_row + 1):
        code_cell = ws.cell(row=row, column=code_col).value
        if code_cell is None or (isinstance(code_cell, str) and not code_cell.strip()):
            # On s'arrête dès la première ligne vide dans la colonne Code DI
            break
        code = str(code_cell).strip()
        entry = sensis.get(code)
        if entry is None:
            missing_codes.append(code)
            continue

        notional = ws.cell(row=row, column=col_notional).value

        def set_value(letter: str, value: object) -> None:
            idx = letter_to_index(letter) + 1
            ws.cell(row=row, column=idx).value = value

        dirty_pct = entry.dirty_pct
        clean_pct = entry.clean_pct
        accrued_pct = entry.accrued_pct

        set_value("BE", dirty_pct)
        set_value("BD", _mul(dirty_pct, notional))
        set_value("BG", clean_pct)
        set_value("BF", _mul(clean_pct, notional))
        set_value("BI", accrued_pct)
        set_value("BH", _mul(accrued_pct, notional))

        set_value("T", entry.duration_leg1)
        set_value("U", entry.sensis_leg1)
        set_value("AK", entry.duration_leg2)
        set_value("AL", entry.sensis_leg2)
        set_value("AT", entry.duration_total)
        set_value("AU", _sum_optional([entry.sensis_leg1, entry.sensis_leg2]))

        # Comparatifs CTP/Bloomberg
        def get_value(letter: str) -> object:
            idx = letter_to_index(letter) + 1
            return ws.cell(row=row, column=idx).value

        set_value("BK", _sub_optional(get_value("AN"), get_value("AW")))
        set_value("BL", _sub_optional(get_value("AO"), get_value("AX")))
        set_value("BS", _sub_optional(get_value("AN"), get_value("BD")))
        set_value("BT", _sub_optional(get_value("AO"), get_value("BE")))
        set_value("BU", _sub_optional(get_value("AP"), get_value("BF")))
        set_value("BV", _sub_optional(get_value("AQ"), get_value("BG")))
        set_value("BW", _sub_optional(get_value("AR"), get_value("BH")))
        set_value("BX", _sub_optional(get_value("AS"), get_value("BI")))
        set_value("CA", _sub_optional(get_value("AW"), get_value("BD")))
        set_value("CB", _sub_optional(get_value("AX"), get_value("BE")))

        updated += 1
        updated_rows.append(
            {
                "Code DI": code,
                "Dirty Price (%)": dirty_pct,
                "Clean Price (%)": clean_pct,
                "Couru (%)": accrued_pct,
                "Sensis L1": entry.sensis_leg1,
                "Sensis L2": entry.sensis_leg2,
                "Duration L1": entry.duration_leg1,
                "Duration L2": entry.duration_leg2,
                "Duration Totale": entry.duration_total,
            }
        )

    wb.save(workbook_path)
    return updated, missing_codes, updated_rows