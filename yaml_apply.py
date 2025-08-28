from __future__ import annotations
from dataclasses import dataclass
from typing import Any, Dict, Iterable
from pathlib import Path
import pandas as pd
import numpy as np
import yaml
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

from excel_read import _norm, get_col
from template_write import build_targets_index, letter_to_index
from data_rules import valid_code_mask

__all__ = [
    "load_cfg", "preview_yaml_rows", "integrate_yaml_to_template",
]

# Ciblage de types (casting) basé sur le libellé cible
NUMERIC_TARGETS = { _norm(s) for s in [
    "Notional","Dirty Value","Clean Value","Accrued Interest","Spread (bp)",
    "Dirty Value(%)","Clean Value(%)","Accrued Interest(%)",
]}
PERCENT_TARGETS = { _norm(s) for s in [
    "Dirty Value(%)","Clean Value(%)","Accrued Interest(%)",
] }
DATE_TARGETS = { _norm(s) for s in ["Start Date","End Date"] }


def load_cfg(yaml_text: str) -> dict:
    cfg = yaml.safe_load(yaml_text) or {}
    if not isinstance(cfg, dict):
        raise ValueError("YAML invalide: racine doit être un mapping")
    return cfg


def _lookup_source_name(df: pd.DataFrame, logical: str) -> str | None:
    """Retourne la vraie colonne (sensible aux variations casse/espaces), sinon None."""
    try:
        return get_col(df, logical, required=False)
    except Exception:
        return None


def _value_from_source_spec(row: pd.Series, orders_by_file: dict[str, dict[str, str]], df_all: pd.DataFrame, spec: dict) -> Any:
    if spec is None:
        return None
    # 1) source_letter -> index dans l'ordre de colonnes du fichier d'origine
    if "source_letter" in spec and spec["source_letter"]:
        file = row.get("__source_file__")
        letter = str(spec["source_letter"]).upper()
        letter_map = orders_by_file.get(file, {}) or {}
        colname = letter_map.get(letter)
        if colname:
            return row.get(colname)
        return None
    # 2) source -> lookup tolérant sur df_all
    if "source" in spec and spec["source"]:
        col = _lookup_source_name(df_all, str(spec["source"]))
        if col is not None:
            return row.get(col)
        return None
    # 3) source_leg (base + leg)
    if "source_leg" in spec:
        base = spec["source_leg"].get("base")
        leg  = int(spec["source_leg"].get("leg", 1))
        # Trouver la première colonne qui matche base et leg via normalisation simple
        target = _norm(base)
        for c in df_all.columns:
            n = _norm(str(c))
            if n.startswith(target) and f"(leg{leg})" in n:
                return row.get(c)
        # fallback: base simple
        col = _lookup_source_name(df_all, base)
        return row.get(col) if col else None
    return None


def _parse_number(x: Any) -> float | None:
    try:
        if x is None or (isinstance(x, float) and np.isnan(x)):
            return None
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip()
        if s == "":
            return None
        s = s.replace("\u00a0", " ")  # nbsp
        s = s.replace(" ", "")
        s = s.replace("'", "")
        s = s.replace(",", ".")
        return float(s)
    except Exception:
        return None


def _parse_date(x: Any) -> datetime | None:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    if isinstance(x, datetime):
        return x
    # Excel serial (rough): base 1899-12-30
    if isinstance(x, (int, float)) and x > 59:
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(x))
        except Exception:
            pass
    s = str(x).strip()
    if not s:
        return None
    # Try dayfirst then monthfirst
    for dayfirst in (True, False):
        try:
            dt = pd.to_datetime(s, dayfirst=dayfirst, errors="raise")
            if isinstance(dt, pd.Timestamp):
                return dt.to_pydatetime()
            return dt
        except Exception:
            continue
    return None


def _cast_for_target(target_label: str, value: Any) -> Any:
    k = _norm(target_label) if target_label else ""
    if k in NUMERIC_TARGETS or k in PERCENT_TARGETS:
        return _parse_number(value)
    if k in DATE_TARGETS:
        return _parse_date(value)
    return value

def _apply_number_format(cell, target_label: str | None) -> None:
    if not target_label:
        return
    k = _norm(target_label)
    if k in PERCENT_TARGETS:
        cell.number_format = "0.00%"  # ex: 1.002 -> 100.2%
    elif k in DATE_TARGETS:
        cell.number_format = "dd/mm/yyyy"  # ex: 28/08/2025


def _eval_expr(expr: str, env: dict[str, float | None]) -> float | None:
    safe_env = {k: (float(v) if v is not None else None) for k, v in env.items()}
    def _div(a, b):
        if a is None or b in (None, 0):
            return None
        return a / b
    allowed: dict[str, Any] = {**safe_env, "_div": _div, "None": None}
    try:
        return eval(expr, {"__builtins__": {}}, allowed)
    except Exception:
        return None


def preview_yaml_rows(df_all: pd.DataFrame, orders_by_file: dict[str, list[str]], cfg: dict, limit: int = 200) -> pd.DataFrame:
    data = df_all.copy()
    # Filtre Code DI si présent
    c_code = get_col(data, "Custom Attribute5 Value", required=False)
    if c_code:
        data = data[valid_code_mask(data[c_code])].copy()
    var_specs: dict = cfg.get("variables", {}) or {}

    rows_out: list[dict] = []
    for _, row in data.iterrows():
        env: dict[str, float | None] = {}
        for vname, vspec in var_specs.items():
            env[vname] = _parse_number(_value_from_source_spec(row, orders_by_file, df_all, vspec))

        direct_vals = []
        for item in cfg.get("columns", []) or []:
            target = item.get("target")
            tocc = int(item.get("target_occurrence", 1))
            sval = _value_from_source_spec(row, orders_by_file, df_all, item)
            sval = _cast_for_target(target, sval)
            direct_vals.append((target, tocc, sval))

        computed_vals = []
        for item in cfg.get("computed", []) or []:
            target = item.get("target")
            tocc = int(item.get("target_occurrence", 1))
            expr = item.get("expr", "")
            val = _eval_expr(expr, env)
            val = _cast_for_target(target, val)
            computed_vals.append((target, tocc, val))

        base = {"__source_file__": row.get("__source_file__"), "Code DI": row.get(c_code) if c_code else None}
        for k, v in env.items():
            base[f"var:{k}"] = v
        for (t, occ, v) in direct_vals:
            base[f"direct:{t}#{occ}"] = v
        for (t, occ, v) in computed_vals:
            base[f"calc:{t}#{occ}"] = v
        rows_out.append(base)
        if len(rows_out) >= limit:
            break
    return pd.DataFrame(rows_out)


from datetime import date as _date_type

def integrate_yaml_to_template(df_all: pd.DataFrame, orders_by_file: dict[str, list[str]], cfg: dict,
                               template_path: Path, dest_dir: Path, file_tag: str, mode: str,
                               ifts_date: _date_type) -> Path:
    from template_write import copy_template_to_dest

    out_path = copy_template_to_dest(template_path, dest_dir, file_tag, mode)
    keep_vba = (out_path.suffix.lower() == ".xlsm")
    wb = load_workbook(out_path, keep_vba=keep_vba, data_only=False)

    ws_name = cfg.get("sheet")
    ws = wb[ws_name] if ws_name in wb.sheetnames else wb.active
    # Écrire la date d'IFTs en B3
    b3 = ws["B3"]
    b3.value = ifts_date
    b3.number_format = "dd/mm/yyyy"

    header_row = int(cfg.get("header_row", 6))
    targets_index = build_targets_index(ws, header_row)

    start_row = header_row + 1
    var_specs: dict = cfg.get("variables", {}) or {}

    # Filtre Code DI si présent
    data = df_all.copy()
    c_code = get_col(data, "Custom Attribute5 Value", required=False)
    if c_code:
        data = data[valid_code_mask(data[c_code])].copy()

    rwrite = 0
    for _, row in data.iterrows():
        r = start_row + rwrite
        # (a) Direct
        for item in cfg.get("columns", []) or []:
            cidx = None
            if item.get("target_letter"):
                cidx = column_index_from_string(item["target_letter"])
            elif item.get("target"):
                lab = _norm(item["target"])
                occ = int(item.get("target_occurrence", 1))
                col_list = targets_index.get(lab, [])
                if 1 <= occ <= len(col_list):
                    cidx = col_list[occ-1]
            if cidx is None:
                continue
            sval = _value_from_source_spec(row, orders_by_file, df_all, item)
            sval = _cast_for_target(item.get("target"), sval)
            cell = ws.cell(row=r, column=cidx)
            cell.value = sval
            _apply_number_format(cell, item.get("target"))

        # (b) Env vars
        env: dict[str, float | None] = {}
        for vname, vspec in var_specs.items():
            env[vname] = _parse_number(_value_from_source_spec(row, orders_by_file, df_all, vspec))

        # (c) Computed
        for item in cfg.get("computed", []) or []:
            cidx = None
            if item.get("target_letter"):
                cidx = column_index_from_string(item["target_letter"])
            elif item.get("target"):
                lab = _norm(item["target"])
                occ = int(item.get("target_occurrence", 1))
                col_list = targets_index.get(lab, [])
                if 1 <= occ <= len(col_list):
                    cidx = col_list[occ-1]
            if cidx is None:
                continue
            val = _eval_expr(item.get("expr", ""), env)
            val = _cast_for_target(item.get("target"), val)
            cell = ws.cell(row=r, column=cidx)
            cell.value = val
            _apply_number_format(cell, item.get("target"))

        rwrite += 1

    wb.save(out_path)
    return out_path
